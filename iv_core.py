"""
iv_core.py — shared logic for the Intrinsic Value Calculator tools.

Used by both:
  - iv_calculator_gui.py   (desktop GUI, Windows/Mac/Linux, tkinter)
  - web/app.py             (Flask backend for the online one-page portal)

Everything here is pure data-fetch + openpyxl-write logic, no UI code.
"""

import datetime
import warnings

import openpyxl
import yfinance as yf

warnings.filterwarnings("ignore")

# ----------------------------------------------------------------------
# Reference data
# ----------------------------------------------------------------------

CURRENCY_SYMBOL_MAP = {
    "AUD": "AUD$", "GBP": "GBP$", "GBp": "GBP$", "CNY": "RMB$", "RMB": "RMB$",
    "EUR": "EUR$", "HKD": "HKD$", "IDR": "IDR$", "MYR": "RM$", "SGD": "SGD$", "USD": "USD$",
}

# Long-term GDP growth + 1% used for Yr 11-20 FCF growth (20y tab only).
GDP_GROWTH_PLUS_1_BY_COUNTRY = {
    "United States": 0.0418,
    "China": 0.07,
    "Hong Kong": 0.07,
}
DEFAULT_GDP_GROWTH_PLUS_1 = 0.0418  # fallback: US rate, flagged for review

# Discount-rate model (Risk Free Rate + Beta x Market Risk Premium), matching
# the "Discount Rate for US Stocks" / "...China/HK Stocks" tables built into
# the template.
DISCOUNT_MODEL_BY_REGION = {
    "US": {"risk_free": 0.016, "market_premium": 0.0428},
    "CHINA_HK": {"risk_free": 0.013, "market_premium": 0.066},
}
CHINA_HK_COUNTRIES = {"China", "Hong Kong"}

TEMPLATE_SHEETS = {20: "VMI (20y)", 10: "VMI (10y)"}


# ----------------------------------------------------------------------
# Small helpers for digging values out of yfinance DataFrames
# ----------------------------------------------------------------------

def _find_row(df, candidates):
    if df is None or df.empty:
        return None
    lower_index = {str(idx).strip().lower(): idx for idx in df.index}
    for cand in candidates:
        key = cand.strip().lower()
        if key in lower_index:
            return df.loc[lower_index[key]]
    return None


def _latest(row):
    if row is None:
        return None
    for val in row.tolist():
        if val is not None and val == val:
            return float(val)
    return None


def _sum_last4(row):
    if row is None:
        return None
    vals = [v for v in row.tolist() if v is not None and v == v]
    if not vals:
        return None
    return float(sum(vals[:4]))


def millions(x):
    return None if x is None else x / 1_000_000.0


def default_year():
    return datetime.datetime.now().year


# ----------------------------------------------------------------------
# Discount rate (Risk Free + Beta x Market Risk Premium)
# ----------------------------------------------------------------------

def compute_discount_rate(beta, country):
    region = "CHINA_HK" if country in CHINA_HK_COUNTRIES else "US"
    model = DISCOUNT_MODEL_BY_REGION[region]
    if beta is None:
        beta = 1.0  # neutral fallback, flagged by caller
    beta_clamped = min(max(beta, 0.8), 1.6)
    rate = model["risk_free"] + beta_clamped * model["market_premium"]
    return rate, region, beta_clamped


# ----------------------------------------------------------------------
# Growth rate fetch
# ----------------------------------------------------------------------

def fetch_growth_yr1_5(t, income_annual):
    try:
        ge = t.growth_estimates
        if ge is not None and not ge.empty:
            for idx_label in ge.index:
                low = str(idx_label).lower()
                if "5 year" in low and "past" not in low:
                    for col in ge.columns:
                        val = ge.loc[idx_label, col]
                        if val is not None and val == val:
                            val = float(val)
                            if abs(val) > 1.5:
                                val = val / 100.0
                            return val, f"Analyst LT growth estimate ({idx_label}, {col})"
    except Exception:
        pass

    try:
        eps_row = _find_row(income_annual, ["Diluted EPS", "Basic EPS"])
        if eps_row is not None:
            vals = [v for v in eps_row.tolist() if v is not None and v == v]
            if len(vals) >= 2 and vals[-1] > 0 and vals[0] > 0:
                n_periods = len(vals) - 1
                cagr = (vals[0] / vals[-1]) ** (1 / n_periods) - 1
                return float(cagr), f"Historical EPS CAGR over {n_periods} yrs"
    except Exception:
        pass

    try:
        eg = t.info.get("earningsGrowth")
        if eg is not None:
            return float(eg), "info['earningsGrowth'] (trailing, weak fallback)"
    except Exception:
        pass

    return None, "NOT FOUND - please enter manually"


# ----------------------------------------------------------------------
# Main data fetch
# ----------------------------------------------------------------------

def fetch_stock_data(ticker, manual_last_close=None, manual_discount_rate=None):
    """Pull everything needed from Yahoo Finance for one ticker.

    manual_last_close / manual_discount_rate let a caller override the
    auto-fetched last close price / discount rate; leave as None to use
    the automatically fetched / computed value.
    """
    t = yf.Ticker(ticker)
    info = t.info or {}

    # --- Cash flow (TTM via last 4 quarters) ---
    q_cf = t.quarterly_cash_flow
    ocf_row = _find_row(q_cf, ["Operating Cash Flow", "Total Cash From Operating Activities"])
    fcf_row = _find_row(q_cf, ["Free Cash Flow"])
    capex_row = _find_row(q_cf, ["Capital Expenditure", "Capital Expenditures"])

    ocf_ttm = _sum_last4(ocf_row)
    fcf_ttm = _sum_last4(fcf_row)
    if fcf_ttm is None and ocf_ttm is not None and capex_row is not None:
        capex_ttm = _sum_last4(capex_row)
        fcf_ttm = ocf_ttm + capex_ttm

    # --- Net income (latest annual) ---
    income_annual = t.income_stmt
    ni_row = _find_row(income_annual, ["Net Income", "Net Income Common Stockholders"])
    net_income_annual = _latest(ni_row)

    # --- Free Cash Flow (Current) per the 1.5x rule ---
    if ocf_ttm is not None and net_income_annual not in (None, 0) and ocf_ttm > 1.5 * net_income_annual:
        method = "Free Cash Flow (Current)"
        fcf_current_value = fcf_ttm
        rule_note = f"OCF TTM ({ocf_ttm:,.0f}) > 1.5x Net Income ({net_income_annual:,.0f}) -> used Free Cash Flow TTM"
    else:
        method = "Net Income (Current)"
        fcf_current_value = net_income_annual
        rule_note = (
            f"OCF TTM ({ocf_ttm if ocf_ttm is not None else 'n/a'}) did NOT exceed "
            f"1.5x Net Income ({net_income_annual if net_income_annual is not None else 'n/a'}) "
            f"-> used Net Income (latest FY)"
        )

    # --- Balance sheet (latest quarter) ---
    q_bs = t.quarterly_balance_sheet
    total_debt = _latest(_find_row(q_bs, ["Total Debt"]))
    if total_debt is None:
        st_debt = _latest(_find_row(q_bs, ["Current Debt", "Short Long Term Debt"]))
        lt_debt = _latest(_find_row(q_bs, ["Long Term Debt"]))
        total_debt = (st_debt or 0) + (lt_debt or 0)

    cash_sti = _latest(_find_row(q_bs, ["Cash Cash Equivalents And Short Term Investments"]))
    if cash_sti is None:
        cash = _latest(_find_row(q_bs, ["Cash And Cash Equivalents"]))
        sti = _latest(_find_row(q_bs, ["Other Short Term Investments"]))
        cash_sti = (cash or 0) + (sti or 0)

    # --- Growth rates ---
    growth_yr1_5, growth_source = fetch_growth_yr1_5(t, income_annual)
    growth_yr6_10 = min(0.15, growth_yr1_5 / 2) if growth_yr1_5 is not None else None

    country = info.get("country", "United States")
    growth_yr11_20 = GDP_GROWTH_PLUS_1_BY_COUNTRY.get(country, DEFAULT_GDP_GROWTH_PLUS_1)
    growth_yr11_20_flagged = country not in GDP_GROWTH_PLUS_1_BY_COUNTRY

    # --- Shares outstanding ---
    shares_out = info.get("sharesOutstanding")

    # --- Currency / FX ---
    stmt_ccy = (info.get("financialCurrency") or "USD").upper()
    stock_ccy = (info.get("currency") or "USD").upper()
    fx_rate = 1.0
    if stmt_ccy != stock_ccy:
        try:
            pair = yf.Ticker(f"{stmt_ccy}{stock_ccy}=X")
            fx_hist = pair.history(period="5d")
            if not fx_hist.empty:
                fx_rate = float(fx_hist["Close"].iloc[-1])
        except Exception:
            pass

    # --- Last close price ---
    last_close = manual_last_close
    last_close_source = "Manual override"
    if last_close is None:
        try:
            last_close = float(t.fast_info["last_price"])
            last_close_source = "Yahoo Finance (fast_info last_price)"
        except Exception:
            try:
                hist = t.history(period="5d")
                last_close = float(hist["Close"].iloc[-1])
                last_close_source = "Yahoo Finance (last 5d history)"
            except Exception:
                last_close = None
                last_close_source = "NOT FOUND - please enter manually"

    # --- Discount rate ---
    discount_rate = manual_discount_rate
    discount_rate_source = "Manual override"
    if discount_rate is None:
        beta = info.get("beta")
        discount_rate, region, beta_used = compute_discount_rate(beta, country)
        discount_rate_source = (
            f"Risk Free + Beta x Market Premium ({region}), beta={beta if beta is not None else 'n/a (assumed 1.0)'}"
            f"{' clamped to ' + str(beta_used) if beta is not None and beta_used != beta else ''}"
        )

    return {
        "ticker": ticker,
        "name": info.get("longName") or info.get("shortName") or ticker,
        "method": method,
        "fcf_current_value_m": millions(fcf_current_value),
        "total_debt_m": millions(total_debt),
        "cash_sti_m": millions(cash_sti),
        "growth_yr1_5": growth_yr1_5,
        "growth_yr1_5_source": growth_source,
        "growth_yr6_10": growth_yr6_10,
        "growth_yr11_20": growth_yr11_20,
        "growth_yr11_20_flagged": growth_yr11_20_flagged,
        "country": country,
        "shares_out_m": millions(shares_out),
        "stmt_currency_symbol": CURRENCY_SYMBOL_MAP.get(stmt_ccy, stmt_ccy + "$"),
        "stock_currency_symbol": CURRENCY_SYMBOL_MAP.get(stock_ccy, stock_ccy + "$"),
        "fx_rate": fx_rate,
        "rule_note": rule_note,
        "last_close": last_close,
        "last_close_source": last_close_source,
        "discount_rate": discount_rate,
        "discount_rate_source": discount_rate_source,
    }


# ----------------------------------------------------------------------
# Excel writing
# ----------------------------------------------------------------------

def _set_comment(cell, text, author="iv_calculator"):
    from openpyxl.comments import Comment
    cell.comment = Comment(text, author)


def populate_sheet(ws, data, current_year, is_20y):
    ws["F8"] = data["name"]
    ws["F10"] = data["ticker"]

    ws["C12"] = data["method"]
    ws["G12"] = data["fcf_current_value_m"]
    _set_comment(ws["G12"], f"Auto-populated via yfinance. Rule: {data['rule_note']}")

    ws["G14"] = data["total_debt_m"]
    _set_comment(ws["G14"], "Source: Yahoo Finance, latest quarterly balance sheet, Total Debt (ST + LT).")

    ws["G16"] = data["cash_sti_m"]
    _set_comment(ws["G16"], "Source: Yahoo Finance, latest quarterly balance sheet, Cash & Short-Term Investments.")

    if data["growth_yr1_5"] is not None:
        ws["F18"] = data["growth_yr1_5"]
    _set_comment(ws["F18"], f"Source: {data['growth_yr1_5_source']}")

    if data["growth_yr6_10"] is not None:
        ws["F20"] = data["growth_yr6_10"]
    _set_comment(ws["F20"], "Rule: min(15%, half of Yr 1-5 growth rate).")

    if is_20y:
        ws["F22"] = data["growth_yr11_20"]
        note = f"Long-term GDP growth + 1% ({data['country']})."
        if data["growth_yr11_20_flagged"]:
            note += " Country not in the built-in table (US/China/HK) - defaulted to 4.18%, please verify."
        _set_comment(ws["F22"], note)

    if data["shares_out_m"] is not None:
        ws["F26"] = data["shares_out_m"]
    _set_comment(ws["F26"], "Source: Yahoo Finance, sharesOutstanding (converted to millions).")

    ws["S8"] = data["stmt_currency_symbol"]
    ws["S10"] = data["stock_currency_symbol"]
    ws["R16"] = data["fx_rate"]

    if data["last_close"] is not None:
        ws["M22"] = data["last_close"]
    _set_comment(ws["M22"], f"Source: {data['last_close_source']}")

    ws["L24"] = current_year

    ws["L26"] = data["discount_rate"]
    _set_comment(ws["L26"], f"Source: {data['discount_rate_source']}")


def build_workbook(template_path, output_path, ticker, current_year=None,
                    manual_last_close=None, manual_discount_rate=None, log=print):
    current_year = current_year or default_year()
    log(f"Fetching data for {ticker} from Yahoo Finance...")
    data = fetch_stock_data(ticker, manual_last_close, manual_discount_rate)
    log(f"  Name: {data['name']}")
    log(f"  Method selected: {data['method']}  ({data['rule_note']})")
    log(f"  FCF (Current): {data['fcf_current_value_m']}")
    log(f"  Total Debt: {data['total_debt_m']}")
    log(f"  Cash & ST Inv: {data['cash_sti_m']}")
    log(f"  Growth Yr1-5: {data['growth_yr1_5']}  ({data['growth_yr1_5_source']})")
    log(f"  Growth Yr6-10: {data['growth_yr6_10']}")
    log(f"  Growth Yr11-20: {data['growth_yr11_20']}")
    log(f"  Shares Out (m): {data['shares_out_m']}")
    log(f"  Last Close: {data['last_close']}  ({data['last_close_source']})")
    log(f"  Discount Rate: {data['discount_rate']}  ({data['discount_rate_source']})")

    wb = openpyxl.load_workbook(template_path)
    populate_sheet(wb[TEMPLATE_SHEETS[20]], data, current_year, is_20y=True)
    populate_sheet(wb[TEMPLATE_SHEETS[10]], data, current_year, is_20y=False)
    wb.save(output_path)
    log(f"\nSaved: {output_path}")
    return data
