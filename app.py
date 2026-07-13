import os, re
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, abort, send_from_directory
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
DATA_DIR = Path(os.getenv("EXCEL_DATA_DIR", Path(__file__).parent / "data"))
ALLOWED = {".xlsx", ".xlsm"}

def clean_ticker(path):
    name = re.sub(r"_Intrinsic_Value$", "", path.stem, flags=re.I)
    return name.upper()

def excel_files():
    DATA_DIR.mkdir(exist_ok=True)
    return sorted([p for p in DATA_DIR.iterdir() if p.is_file() and p.suffix.lower() in ALLOWED and p.name.lower().endswith(("_intrinsic_value.xlsx", "_intrinsic_value.xlsm"))], key=lambda p: clean_ticker(p))

def fmt(value, number_format="General"):
    if value is None: return ""
    if isinstance(value, datetime): return value.strftime("%Y-%m-%d")
    if isinstance(value, bool): return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        nf=(number_format or "").lower()
        if "%" in nf: return f"{value:.2%}"
        if abs(value)>=1_000_000: return f"{value:,.0f}"
        if abs(value)>=1000: return f"{value:,.2f}"
        if abs(value)<1 and value!=0: return f"{value:.4f}".rstrip("0").rstrip(".")
        return f"{value:,.2f}".rstrip("0").rstrip(".")
    return str(value)

def color_hex(color):
    if not color: return None
    if color.type == 'rgb' and color.rgb:
        v=color.rgb[-6:]
        return v if v != '000000' else None
    return None

def workbook_data(path):
    formulas=load_workbook(path, data_only=False, read_only=False)
    values=load_workbook(path, data_only=True, read_only=False)
    sheets=[]
    for ws, vws in zip(formulas.worksheets, values.worksheets):
        merged={}
        skip=set()
        for rng in ws.merged_cells.ranges:
            merged[(rng.min_row,rng.min_col)] = (rng.max_row-rng.min_row+1, rng.max_col-rng.min_col+1)
            for r in range(rng.min_row,rng.max_row+1):
                for c in range(rng.min_col,rng.max_col+1):
                    if (r,c)!=(rng.min_row,rng.min_col): skip.add((r,c))
        rows=[]
        max_col=min(ws.max_column, 45)
        max_row=min(ws.max_row, 250)
        for r in range(1,max_row+1):
            cells=[]
            for c in range(1,max_col+1):
                if (r,c) in skip: continue
                cell=ws.cell(r,c); vcell=vws.cell(r,c)
                raw=vcell.value if vcell.value is not None else cell.value
                if isinstance(raw,str) and raw.startswith('='): raw='Formula not cached in Excel'
                rowspan,colspan=merged.get((r,c),(1,1))
                style=[]
                fill=color_hex(cell.fill.fgColor)
                font=color_hex(cell.font.color) if cell.font and cell.font.color else None
                if fill: style.append(f"background-color:#{fill}")
                if font: style.append(f"color:#{font}")
                if cell.font and cell.font.bold: style.append("font-weight:700")
                if cell.alignment and cell.alignment.horizontal: style.append(f"text-align:{cell.alignment.horizontal}")
                width=ws.column_dimensions[get_column_letter(c)].width
                cells.append({"value":fmt(raw,cell.number_format),"rowspan":rowspan,"colspan":colspan,"style":";".join(style),"width":width})
            if any(x['value'] for x in cells): rows.append(cells)
        sheets.append({"name":ws.title,"rows":rows})
    return sheets, values

def find_value(wb, labels, numeric=False):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                text=str(cell.value or '').strip().lower()
                for label in labels:
                    if label in text:
                        for dc in range(1,6):
                            v=ws.cell(cell.row, cell.column+dc).value
                            if numeric and isinstance(v,(int,float)):
                                return v
                            if not numeric and (isinstance(v,(int,float)) or (isinstance(v,str) and v.strip())):
                                return v
    return None

def summary(wb, ticker):
    fields=[
      ("Company", ["name of stock"], False, "General"),
      ("Last Close", ["last close"], True, "General"),
      ("Intrinsic Value", ["final intrinsic value per share"], True, "General"),
      ("Discount / Premium", ["(discount)/premium"], True, "0.00%"),
      ("Discount Rate", ["discount rate"], True, "0.00%"),
      ("Shares Outstanding", ["no. of shares outstanding"], True, "General"),
    ]
    out=[]
    for label, aliases, numeric, number_format in fields:
        v=find_value(wb, aliases, numeric=numeric)
        if v is not None:
            out.append({"label":label,"value":fmt(v, number_format)})
    return out

@app.route('/')
def index():
    files=excel_files()
    stocks=[{"ticker":clean_ticker(p),"filename":p.name} for p in files]
    if not stocks: return render_template('empty.html')
    return render_template('index.html', stocks=stocks, active=stocks[0]['ticker'])

@app.route('/stock/<ticker>')
def stock(ticker):
    files=excel_files(); lookup={clean_ticker(p):p for p in files}
    key=ticker.upper()
    if key not in lookup: abort(404)
    sheets, wb=workbook_data(lookup[key])
    stocks=[{"ticker":clean_ticker(p),"filename":p.name} for p in files]
    return render_template('stock.html', stocks=stocks, active=key, sheets=sheets, summary=summary(wb,key), filename=lookup[key].name)

@app.route('/download/<ticker>')
def download(ticker):
    lookup={clean_ticker(p):p for p in excel_files()}; p=lookup.get(ticker.upper())
    if not p: abort(404)
    return send_from_directory(DATA_DIR, p.name, as_attachment=True)

@app.errorhandler(404)
def not_found(e): return render_template('404.html'),404

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT',5000)), debug=os.environ.get('FLASK_DEBUG')=='1')
